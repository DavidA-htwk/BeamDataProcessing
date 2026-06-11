"""
Extract values from smoothing log .txt files and write them into Book2.xlsx.

For each scenario column in the workbook, the script:
  1. Derives the expected log file path from the scenario name.
  2. Parses the results_10_CONNECTING_DUCT2 block.
  3. Writes Total power (W) into the "Total power" row and
     Peak density (after) (W/m²) into the "Max, W/m2" row.

Folder layout expected on disk (all under BASE_DIR):
    OUTPUT_HFHC/dnb_3_+10_+2/SMOOTHED/smoothing_log.txt
    OUTPUT_FFFC/dnb_7_-10_-2/SMOOTHED/smoothing_log.txt
    ...

Usage:
    python extract_to_excel.py --base <path_to_output_root> \
                               --xlsx  <path_to_Book2.xlsx>  \
                               [--log-name <filename>]       \
                               [--out   <path_for_result>]

    python Extract_results.py `
    --base "D:\attelnd\Beam_on_target\BeamOnTarget" `
    --xlsx "D:\attelnd\Value_extract\Book2_filled.xlsx" `
    --out  "D:\attelnd\Value_extract\Book2_results.xlsx"
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

TARGET_BLOCK = "results_10_CONNECTING_DUCT2"

TOTAL_POWER_RE  = re.compile(r"Total power\s*:\s*([0-9.eE+\-]+)\s*W")
PEAK_AFTER_RE   = re.compile(r"Peak density \(after\)\s*:\s*([0-9.eE+\-]+)\s*W/m")


def scenario_to_rel_path(scenario: str) -> Path:
    """
    Convert a scenario name such as 'hfhc-3+10+2' to the relative path
    OUTPUT_HFHC/dnb_3_+10_+2
    """
    typ  = scenario[:4].upper()          # HFHC
    rest = scenario[5:]                  # 3+10+2
    m = re.match(r"(\d+)([+-]\d+)([+-]\d+)", rest)
    if not m:
        raise ValueError(f"Cannot parse scenario '{scenario}'")
    n, y, z = m.group(1), m.group(2), m.group(3)
    return Path(f"OUTPUT_{typ}") / f"dnb_{n}_{y}_{z}"


def parse_log(log_path: Path):
    """
    Return (total_power_W, peak_density_after_Wm2) from the
    results_10_CONNECTING_DUCT2 block, or (None, None) if not found.
    """
    try:
        text = log_path.read_text(encoding="utf-8", errors="replace")
    except FileNotFoundError:
        return None, None

    # Find the block for CONNECTING_DUCT2
    # A block starts at "--- results_10_CONNECTING_DUCT2 ---" and ends at
    # the next "--- results_" line or end of file.
    block_start = text.find(f"--- {TARGET_BLOCK}")
    if block_start == -1:
        return None, None

    next_block = text.find("--- results_", block_start + 1)
    block = text[block_start: next_block if next_block != -1 else len(text)]

    m_power = TOTAL_POWER_RE.search(block)
    m_peak  = PEAK_AFTER_RE.search(block)

    total_power   = float(m_power.group(1)) if m_power else None
    peak_density  = float(m_peak.group(1))  if m_peak  else None
    return total_power, peak_density


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--base",     required=True,
                        help="Root directory containing OUTPUT_HFHC, OUTPUT_FFFC, … folders")
    parser.add_argument("--xlsx",     required=True,
                        help="Path to Book2.xlsx (will be read and re-saved)")
    parser.add_argument("--log-name", default="smooth_log.txt",
                        help="Name of the log file inside each SMOOTHED folder "
                             "(default: smooth_log.txt)")
    parser.add_argument("--out",      default=None,
                        help="Output xlsx path (defaults to overwriting --xlsx)")
    args = parser.parse_args()

    base_dir  = Path(args.base)
    xlsx_path = Path(args.xlsx)
    out_path  = Path(args.out) if args.out else xlsx_path

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # -----------------------------------------------------------------------
    # Build column map:  scenario_name -> (total_power_col, max_col)
    #
    # Layout (1-indexed):
    #   Header rows  1 + 2  → Total power (W)
    #   Header rows 16 + 17 → Max, W/m2
    #
    # Row 1  : scenario name  (every column that isn't "Scenario\nPanel")
    # Row 2  : "Total power (W)"  – same column as scenario
    # Row 16 : scenario name  (mirror of row 1)
    # Row 17 : "Max, W/m2"    – same column as scenario
    #
    # Data rows start at row 3 for the power block and row 18 for the max block.
    # There are no data rows yet; we find the first empty data row in each block.
    # -----------------------------------------------------------------------

    # Collect scenario → column index mapping from row 1
    scenario_col = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        val = cell.value
        if val and val != "Scenario\nPanel":
            scenario_col[val] = col_idx

    # Identify the first empty data row for each block
    # Block 1 (Total power): header in rows 1-2, data from row 3
    # Block 2 (Max W/m2):    header in rows 16-17, data from row 18
    # We write into row 3 and row 18 respectively (the first data row).
    POWER_DATA_ROW = 3
    MAX_DATA_ROW   = 18

    hits   = 0
    misses = []

    for scenario, col in scenario_col.items():
        rel = scenario_to_rel_path(scenario)
        log_file = base_dir / rel / "SMOOTHED" / args.log_name

        total_power, peak_density = parse_log(log_file)

        if total_power is None and peak_density is None:
            misses.append((scenario, str(log_file)))
            continue

        if total_power is not None:
            ws.cell(row=POWER_DATA_ROW, column=col).value = total_power

        if peak_density is not None:
            ws.cell(row=MAX_DATA_ROW, column=col).value = peak_density

        hits += 1
        print(f"  [OK] {scenario:20s}  power={total_power}  peak_after={peak_density}")

    wb.save(out_path)

    print(f"\n{'='*60}")
    print(f"Written to : {out_path}")
    print(f"Populated  : {hits} scenario(s)")
    if misses:
        print(f"Missing    : {len(misses)} scenario(s) — log file not found or block absent:")
        for scenario, path in misses:
            print(f"             {scenario:20s}  -> {path}")
    else:
        print("Missing    : none")


if __name__ == "__main__":
    main()